#!/usr/bin/env python3
# Python Excel Text Extractor Service
# Extracts text from Excel files (.xls, .xlsx) using openpyxl and xlrd
import sys
import json
import os

try:
    import openpyxl
    import xlrd
    import pandas as pd
except ImportError as e:
    # When imported as module, don't exit, just raise the error
    if __name__ == "__main__":
        print(json.dumps({"success": False, "error": f"Missing required Python packages: {str(e)}", "text": "", "sheets": 0, "metadata": {}, "word_count": 0, "character_count": 0}))
        sys.exit(1)
    else:
        raise ImportError(f"Missing required Python packages: {str(e)}")

def extract_text_with_openpyxl(filepath):
    """Extract text using openpyxl (for .xlsx files)"""
    try:
        workbook = openpyxl.load_workbook(filepath, data_only=True)
        text = ""
        sheet_data = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_text = ""
            rows_data = []
            
            for row in sheet.iter_rows(values_only=True):
                row_text = ""
                row_data = []
                for cell in row:
                    if cell is not None and str(cell).strip():
                        cell_text = str(cell).strip()
                        row_text += cell_text + " "
                        row_data.append(cell_text)
                
                if row_text.strip():
                    rows_data.append(row_data)
                    sheet_text += row_text + "\n"
            
            if sheet_text.strip():
                sheet_data.append({
                    "sheet_name": sheet_name,
                    "text": sheet_text.strip(),
                    "rows": rows_data
                })
                text += f"--- Sheet: {sheet_name} ---\n{sheet_text}\n\n"
        
        workbook.close()
        return {
            "text": text,
            "sheets": len(workbook.sheetnames),
            "sheet_data": sheet_data,
            "method": "openpyxl",
            "success": True
        }
    except Exception as e:
        return {"text": "", "sheets": 0, "method": "openpyxl", "success": False, "error": str(e)}

def extract_text_with_pandas(filepath):
    """Extract text using pandas (alternative method)"""
    try:
        # Try to read all sheets
        excel_file = pd.ExcelFile(filepath)
        text = ""
        sheet_data = []
        
        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(filepath, sheet_name=sheet_name)
            sheet_text = ""
            
            # Convert DataFrame to text
            for index, row in df.iterrows():
                row_text = " ".join([str(cell) for cell in row if pd.notna(cell) and str(cell).strip()])
                if row_text.strip():
                    sheet_text += row_text + "\n"
            
            if sheet_text.strip():
                sheet_data.append({
                    "sheet_name": sheet_name,
                    "text": sheet_text.strip(),
                    "shape": df.shape
                })
                text += f"--- Sheet: {sheet_name} ---\n{sheet_text}\n\n"
        
        return {
            "text": text,
            "sheets": len(excel_file.sheet_names),
            "sheet_data": sheet_data,
            "method": "pandas",
            "success": True
        }
    except Exception as e:
        return {"text": "", "sheets": 0, "method": "pandas", "success": False, "error": str(e)}

def extract_text_with_xlrd(filepath):
    """Extract text using xlrd (for .xls files)"""
    try:
        workbook = xlrd.open_workbook(filepath)
        text = ""
        sheet_data = []
        
        for sheet_name in workbook.sheet_names():
            sheet = workbook.sheet_by_name(sheet_name)
            sheet_text = ""
            rows_data = []
            
            for row_num in range(sheet.nrows):
                row_text = ""
                row_data = []
                for col_num in range(sheet.ncols):
                    cell = sheet.cell(row_num, col_num)
                    if cell.value and str(cell.value).strip():
                        cell_text = str(cell.value).strip()
                        row_text += cell_text + " "
                        row_data.append(cell_text)
                
                if row_text.strip():
                    rows_data.append(row_data)
                    sheet_text += row_text + "\n"
            
            if sheet_text.strip():
                sheet_data.append({
                    "sheet_name": sheet_name,
                    "text": sheet_text.strip(),
                    "rows": rows_data
                })
                text += f"--- Sheet: {sheet_name} ---\n{sheet_text}\n\n"
        
        return {
            "text": text,
            "sheets": len(workbook.sheet_names()),
            "sheet_data": sheet_data,
            "method": "xlrd",
            "success": True
        }
    except Exception as e:
        return {"text": "", "sheets": 0, "method": "xlrd", "success": False, "error": str(e)}

def extract_metadata_from_excel(filepath):
    """Extract metadata from Excel file"""
    try:
        # Try with openpyxl first (for .xlsx)
        try:
            workbook = openpyxl.load_workbook(filepath)
            metadata = {
                "file_type": "xlsx",
                "sheet_count": len(workbook.sheetnames),
                "sheet_names": workbook.sheetnames
            }
            workbook.close()
            return metadata
        except:
            # Try with xlrd (for .xls)
            workbook = xlrd.open_workbook(filepath)
            metadata = {
                "file_type": "xls",
                "sheet_count": len(workbook.sheet_names()),
                "sheet_names": workbook.sheet_names()
            }
            return metadata
    except Exception as e:
        return {"error": str(e)}

def extract_text_from_excel(filepath):
    """Extract text from Excel file using multiple methods"""
    results = []
    
    # Try different extraction methods based on file extension
    file_ext = os.path.splitext(filepath)[1].lower()
    
    if file_ext == '.xlsx':
        methods = [extract_text_with_openpyxl, extract_text_with_pandas]
    else:  # .xls
        methods = [extract_text_with_xlrd, extract_text_with_pandas]
    
    for method in methods:
        result = method(filepath)
        results.append(result)
        
        # If we got good text, use it
        if result["success"] and result["text"].strip() and len(result["text"].strip()) > 10:
            break
    
    # Choose the best result
    successful_results = [r for r in results if r["success"]]
    if not successful_results:
        return {
            "success": False,
            "text": "",
            "sheets": 0,
            "metadata": {},
            "word_count": 0,
            "character_count": 0,
            "error": "All extraction methods failed"
        }
    
    best_result = max(successful_results, key=lambda x: len(x["text"].strip()))
    
    # Get metadata
    metadata = extract_metadata_from_excel(filepath)
    
    # Count words and characters
    text = best_result["text"]
    word_count = len(text.split())
    character_count = len(text)
    
    return {
        "success": True,
        "text": text,
        "sheets": best_result["sheets"],
        "sheet_data": best_result.get("sheet_data", []),
        "metadata": metadata,
        "word_count": word_count,
        "character_count": character_count,
        "extraction_method": best_result["method"],
        "all_methods": [{"method": r["method"], "success": r["success"], "text_length": len(r["text"]), "error": r.get("error")} for r in results]
    }

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps({"success": False, "error": "No file path provided", "text": "", "sheets": 0, "metadata": {}, "word_count": 0, "character_count": 0}))
        sys.exit(1)
    
    excel_filepath = sys.argv[1]
    if not os.path.exists(excel_filepath):
        print(json.dumps({"success": False, "error": f"File not found: {excel_filepath}", "text": "", "sheets": 0, "metadata": {}, "word_count": 0, "character_count": 0}))
        sys.exit(1)
    
    result = extract_text_from_excel(excel_filepath)
    print(json.dumps(result, indent=2))
